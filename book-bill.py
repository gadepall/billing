#Program to generate a bill based on excel input
#Input abcd.xlsx
#Bill Details table.xlsx
#main.tex bill template

#Code by G V V Sharma
#October 11 2024

import openpyxl

#if using termux
import subprocess
import shlex
#end if

# Load the source workbook and select the active sheet
source_wb = openpyxl.load_workbook('tables/abcd.xlsx')
source_sheet = source_wb.active

# Load the destination workbook and select the active sheet
destination_wb = openpyxl.load_workbook('tables/table.xlsx')
destination_sheet = destination_wb.active
for i in range(2, 6):

    #Name
    # Read the value from row i, column 2
    value = source_sheet.cell(row=i, column=2).value
    # Write the value to row 4, column 2
    destination_sheet.cell(row=4, column=2).value = value
    print(value)
    
    #Date
    # Read the value from row i, column 3
    value = source_sheet.cell(row=i, column=3).value
    # Write the value to row 4, column 2
    destination_sheet.cell(row=2, column=5).value = value
    destination_sheet.cell(row=10, column=2).value = value
    print(value)
    
    #Transaction Number
    # Read the value from row i, column 4
    value = source_sheet.cell(row=i, column=4).value
    # Write the value to row 10, column 1
    destination_sheet.cell(row=10, column=1).value = value
    print(value)
    
    #Bill No
    # Original string
    original_string = "ABCD000"
    
    # Extract the numeric part
    numeric_part = int(original_string[3:])
    
    # Increment the numeric part by 1
    incremented_numeric_part = numeric_part + i
    
    # Combine the prefix with the incremented numeric part
    value = original_string[:3] + str(incremented_numeric_part)
    
    # Write the value to row 3, column 5
    destination_sheet.cell(row=3, column=5).value = value
    
    pdfext = ".pdf"
    pdfext= value + pdfext
    
    print(value, pdfext)
    
    # Save the destination workbook
    destination_wb.save('tables/table.xlsx')
    
    #Executing shell commands
    command = f"ssconvert --export-type=Gnumeric_html:latex tables/table.xlsx tables/table.tex && texfot pdflatex main.tex && mv main.pdf {pdfext}"
    result = subprocess.run(command, shell=True, capture_output=True, text=True)
    
print("Value copied successfully!")

